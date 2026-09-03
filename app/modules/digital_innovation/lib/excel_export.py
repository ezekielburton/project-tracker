# Cost-ledger Excel export: turn a project's cost_summary() (lib/costs.py) into
# an .xlsx workbook. Kept separate from routes/costs.py, which just streams what
# this returns.

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.digital_innovation.lib.costs import DI_COST_TYPE_LABELS

_HEADER_FILL = PatternFill(start_color='4A4A4A', end_color='4A4A4A', fill_type='solid')


def build_cost_ledger_workbook(di_project, summary):
    """Build the workbook in memory and return a BytesIO at the start, ready for
    send_file. `summary` is lib.costs.cost_summary(di_project)'s result, passed
    in so this doesn't re-query."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cost Ledger'

    bold = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF')

    ws['A1'] = di_project.name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Cost ledger — exported {datetime.utcnow().strftime("%d %b %Y")}'
    ws['A2'].font = Font(italic=True, color='666666')

    headers = ['Date', 'Type', 'Feature', 'Description', 'Hours', 'Amount']
    header_row = 4
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left')
        cell.fill = _HEADER_FILL

    row = header_row + 1
    # On-screen ledger is newest-first; the export reverses to oldest-first,
    # which reads more naturally as a log.
    for entry in reversed(summary['entries']):
        ws.cell(row=row, column=1, value=entry.date)
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=2, value=DI_COST_TYPE_LABELS.get(entry.type, entry.type))
        ws.cell(row=row, column=3, value=entry.feature.name if entry.feature else '')
        ws.cell(row=row, column=4, value=entry.description or '')
        ws.cell(row=row, column=5, value=entry.hours if entry.hours else None)
        ws.cell(row=row, column=6, value=entry.amount)
        row += 1

    totals_row = row + 1
    ws.cell(row=totals_row, column=5, value='Total cost').font = bold
    ws.cell(row=totals_row, column=6, value=summary['total_cost']).font = bold

    if summary['client_charge'] is not None:
        ws.cell(row=totals_row + 1, column=5, value='Client charge').font = bold
        ws.cell(row=totals_row + 1, column=6, value=summary['client_charge']).font = bold
        ws.cell(row=totals_row + 2, column=5, value='Projected profit').font = bold
        ws.cell(row=totals_row + 2, column=6, value=summary['projected_profit']).font = bold

    widths = [12, 12, 22, 34, 9, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_performance_workbook(rollup, currency):
    """Performance export — one row per project in the period's rollup, same
    columns as the on-screen table, plus the three summary figures. `rollup` is
    lib.snapshots.get_period_rollup()'s result, passed in so this doesn't
    re-query."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Performance'

    bold = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF')

    ws['A1'] = f"Digital Innovation Performance — {rollup['period_label']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Exported {datetime.utcnow().strftime("%d %b %Y")}'
    ws['A2'].font = Font(italic=True, color='666666')

    ws['A4'] = 'Total cost'
    ws['A4'].font = bold
    ws['B4'] = rollup['total_cost']
    ws['A5'] = 'Closed profit'
    ws['A5'].font = bold
    ws['B5'] = rollup['closed_profit']
    ws['A6'] = 'Projected profit'
    ws['A6'].font = bold
    ws['B6'] = rollup['projected_profit']

    headers = ['Project', 'Status', 'Dev hours', 'Cost', 'Charge', 'Profit']
    header_row = 8
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left')
        cell.fill = _HEADER_FILL

    row = header_row + 1
    for project in rollup['projects']:
        is_closed = project['lifecycle'] in ('closed', 'archived')
        ws.cell(row=row, column=1, value=project['name'])
        ws.cell(row=row, column=2, value='Closed' if is_closed else 'In progress')
        ws.cell(row=row, column=3, value=project['dev_hours'])
        ws.cell(row=row, column=4, value=project['total_cost'])
        ws.cell(row=row, column=5, value=project['client_charge'])
        ws.cell(row=row, column=6, value=project['profit'])
        row += 1

    if not rollup['projects']:
        ws.cell(row=row, column=1, value='No projects in this period.')

    for col in (4, 5, 6):
        for r in range(header_row + 1, row):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None:
                cell.number_format = f'"{currency}" #,##0'
    ws['B4'].number_format = f'"{currency}" #,##0'
    ws['B5'].number_format = f'"{currency}" #,##0'
    ws['B6'].number_format = f'"{currency}" #,##0'

    widths = [26, 12, 11, 14, 14, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
